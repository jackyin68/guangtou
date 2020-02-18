import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook('filename')
    sheet = wb['Sheet1']


    for row in range(2, sheet.max_row + 1):#print(row)确认
        cell = sheet.cell(row, 3)#print(cell.value)确认
        corrected_price = cell.value * 0.9#这里不能直接用cell乘，需要value
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price


    values = Reference(sheet, min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)#为了防止程序有bug覆盖了原文件

