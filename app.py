import openpyxl as xl #不是必须，但可以让代码更简洁
from openpyxl.chart import BarChart, Reference#Reference选择值的范围
wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
#定位到单元格
cell = sheet['a1'] #sheet.cell(1, 1) 一样的效果
#cell的值
print(cell.value)
#了解表格的情况 print(sheet.max_row)
for row in range(2, sheet.max_row + 1):#print(row)确认
    cell = sheet.cell(row, 3)#print(cell.value)确认
    corrected_price = cell.value * 0.9#这里不能直接用cell乘，需要value
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price
#values里就有corrected的price了
values = Reference(sheet, min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('transactions2.xlsx')#为了防止程序有bug覆盖了原文件

